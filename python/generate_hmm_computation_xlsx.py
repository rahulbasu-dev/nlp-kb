"""Generate an Excel workbook showing HMM Forward-Backward-Baum-Welch computations.

Produces: docs/HMM_Forward_Backward_Computation.xlsx
"""

import os
import math
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# HMM parameters
# ---------------------------------------------------------------------------
STATES = ["Rainy", "Sunny"]
OBS_SYMBOLS = ["Umbrella", "No-umbrella"]
OBS_SEQ = [0, 1, 0]  # U, N, U
OBS_LABELS = ["U", "N", "U"]

A = np.array([[0.7, 0.3],
              [0.4, 0.6]])  # transition
B = np.array([[0.9, 0.1],
              [0.2, 0.8]])  # emission  rows=states, cols=obs
PI = np.array([0.5, 0.5])

N_STATES = 2
T = len(OBS_SEQ)
BW_ITERATIONS = 10

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
DARK_BLUE = PatternFill("solid", fgColor="1F4E79")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
LIGHT_GRAY = PatternFill("solid", fgColor="F2F2F2")
WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True, size=11)
ITALIC_GRAY = Font(italic=True, color="808080", size=10)
NORMAL = Font(size=11)
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM_FMT = "0.000000"
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, col, text, span=1, fill=DARK_BLUE, font=WHITE_FONT):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = BORDER
    if span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + span - 1
        )
        for c in range(col + 1, col + span):
            cl = ws.cell(row=row, column=c)
            cl.fill = fill
            cl.font = font
            cl.border = BORDER


def style_subheader(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = LIGHT_BLUE
    cell.font = BOLD
    cell.alignment = CENTER
    cell.border = BORDER


def style_formula(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = ITALIC_GRAY
    cell.alignment = LEFT
    cell.border = BORDER


def write_val(ws, row, col, val, fmt=NUM_FMT, bold=False):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = fmt
    cell.font = Font(bold=bold, size=11)
    cell.alignment = CENTER
    cell.border = BORDER


def write_text(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = CENTER
    cell.border = BORDER


def auto_width(ws, min_width=10, max_width=22):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        best = min_width
        for cell in col_cells:
            if cell.value is not None:
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = best


# ---------------------------------------------------------------------------
# HMM algorithms
# ---------------------------------------------------------------------------

def forward(pi, a, b, obs):
    T_ = len(obs)
    N = len(pi)
    alpha = np.zeros((T_, N))
    alpha[0] = pi * b[:, obs[0]]
    for t in range(1, T_):
        for j in range(N):
            alpha[t, j] = np.sum(alpha[t - 1] * a[:, j]) * b[j, obs[t]]
    return alpha


def backward(a, b, obs):
    T_ = len(obs)
    N = a.shape[0]
    beta = np.zeros((T_, N))
    beta[T_ - 1] = 1.0
    for t in range(T_ - 2, -1, -1):
        for i in range(N):
            beta[t, i] = np.sum(a[i, :] * b[:, obs[t + 1]] * beta[t + 1])
    return beta


def compute_gamma(alpha, beta):
    g = alpha * beta
    g = g / g.sum(axis=1, keepdims=True)
    return g


def compute_xi(alpha, beta, a, b, obs):
    T_ = len(obs)
    N = a.shape[0]
    xi = np.zeros((T_ - 1, N, N))
    for t in range(T_ - 1):
        denom = 0.0
        for i in range(N):
            for j in range(N):
                xi[t, i, j] = alpha[t, i] * a[i, j] * b[j, obs[t + 1]] * beta[t + 1, j]
                denom += xi[t, i, j]
        xi[t] /= denom
    return xi


def baum_welch_step(pi, a, b, obs):
    alpha = forward(pi, a, b, obs)
    beta = backward(a, b, obs)
    gamma = compute_gamma(alpha, beta)
    xi = compute_xi(alpha, beta, a, b, obs)
    N = a.shape[0]
    M = b.shape[1]
    T_ = len(obs)

    pi_new = gamma[0]
    a_new = np.zeros_like(a)
    for i in range(N):
        for j in range(N):
            a_new[i, j] = xi[:, i, j].sum() / gamma[:-1, i].sum()
    b_new = np.zeros_like(b)
    for j in range(N):
        for k in range(M):
            mask = np.array([1.0 if obs[t] == k else 0.0 for t in range(T_)])
            b_new[j, k] = (gamma[:, j] * mask).sum() / gamma[:, j].sum()

    ll = np.log(alpha[-1].sum())
    return pi_new, a_new, b_new, alpha, beta, gamma, xi, ll


# ---------------------------------------------------------------------------
# Sheet 1: HMM Parameters
# ---------------------------------------------------------------------------

def build_parameters_sheet(wb):
    ws = wb.active
    ws.title = "HMM Parameters"

    style_header(ws, 1, 1, "HMM Parameters — Weather/Umbrella Example", span=5)

    # Initial distribution
    r = 3
    style_header(ws, r, 1, "Initial Distribution π", span=3, fill=DARK_BLUE)
    r += 1
    for i, s in enumerate(STATES):
        style_subheader(ws, r, 1, s)
        write_val(ws, r, 2, PI[i])
        r += 1

    r += 1
    style_header(ws, r, 1, "Transition Matrix A", span=3, fill=DARK_BLUE)
    r += 1
    write_text(ws, r, 1, "From \\ To", bold=True)
    for j, s in enumerate(STATES):
        style_subheader(ws, r, 2 + j, s)
    r += 1
    for i, s in enumerate(STATES):
        write_text(ws, r, 1, s, bold=True)
        for j in range(N_STATES):
            write_val(ws, r, 2 + j, A[i, j])
        r += 1

    r += 1
    style_header(ws, r, 1, "Emission Matrix B", span=3, fill=DARK_BLUE)
    r += 1
    write_text(ws, r, 1, "State \\ Obs", bold=True)
    for k, o in enumerate(OBS_SYMBOLS):
        style_subheader(ws, r, 2 + k, o)
    r += 1
    for i, s in enumerate(STATES):
        write_text(ws, r, 1, s, bold=True)
        for k in range(len(OBS_SYMBOLS)):
            write_val(ws, r, 2 + k, B[i, k])
        r += 1

    r += 1
    style_header(ws, r, 1, "Observation Sequence", span=5, fill=DARK_BLUE)
    r += 1
    for t_idx in range(T):
        write_text(ws, r, 1 + t_idx, f"t={t_idx+1}: {OBS_LABELS[t_idx]}", bold=True)

    auto_width(ws)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 2: Forward-Backward
# ---------------------------------------------------------------------------

def build_forward_backward_sheet(wb):
    ws = wb.create_sheet("Forward-Backward")

    alpha = forward(PI, A, B, OBS_SEQ)
    beta = backward(A, B, OBS_SEQ)
    gamma = compute_gamma(alpha, beta)
    p_obs = alpha[-1].sum()

    # ---- Section A: Parameters (cols 1-7) ----
    style_header(ws, 1, 1, "HMM Parameters", span=4)

    r = 3
    write_text(ws, r, 1, "π", bold=True)
    for i, s in enumerate(STATES):
        write_text(ws, r, 2 + i, s, bold=True)
    r += 1
    write_text(ws, r, 1, "")
    for i in range(N_STATES):
        write_val(ws, r, 2 + i, PI[i])

    r += 2
    write_text(ws, r, 1, "A", bold=True)
    for j, s in enumerate(STATES):
        write_text(ws, r, 2 + j, s, bold=True)
    r += 1
    for i, s in enumerate(STATES):
        write_text(ws, r, 1, s, bold=True)
        for j in range(N_STATES):
            write_val(ws, r, 2 + j, A[i, j])
        r += 1

    r += 1
    write_text(ws, r, 1, "B", bold=True)
    for k, o in enumerate(OBS_SYMBOLS):
        write_text(ws, r, 2 + k, o, bold=True)
    r += 1
    for i, s in enumerate(STATES):
        write_text(ws, r, 1, s, bold=True)
        for k in range(len(OBS_SYMBOLS)):
            write_val(ws, r, 2 + k, B[i, k])
        r += 1

    # ---- Section B: Forward (cols 9-14) ----
    fc = 9  # starting column
    style_header(ws, 1, fc, "Forward Algorithm (α)", span=6)
    # Row 2: step labels
    for t_idx in range(T):
        label = "Step 1: Init" if t_idx == 0 else f"Step {t_idx+1}: Recursion"
        style_header(ws, 2, fc + t_idx * 2, label, span=2, fill=LIGHT_BLUE, font=BOLD)
    # Row 3: observation labels
    for t_idx in range(T):
        obs_sym = OBS_LABELS[t_idx]
        style_subheader(ws, 3, fc + t_idx * 2, f"t={t_idx+1} (o={obs_sym})")
        ws.merge_cells(start_row=3, start_column=fc + t_idx * 2,
                       end_row=3, end_column=fc + t_idx * 2 + 1)
    # Row 4: column headers
    for t_idx in range(T):
        for s_idx, s in enumerate(["R", "S"]):
            style_subheader(ws, 4, fc + t_idx * 2 + s_idx, f"α{t_idx+1}({s})")

    # Row 5: formulas
    formulas_fwd = [
        [f"π(R)·B_R(U)={PI[0]}×{B[0,OBS_SEQ[0]]}",
         f"π(S)·B_S(U)={PI[1]}×{B[1,OBS_SEQ[0]]}"],
        [f"[α1(R)·a_RR + α1(S)·a_SR]·B_R(N)",
         f"[α1(R)·a_RS + α1(S)·a_SS]·B_S(N)"],
        [f"[α2(R)·a_RR + α2(S)·a_SR]·B_R(U)",
         f"[α2(R)·a_RS + α2(S)·a_SS]·B_S(U)"],
    ]
    for t_idx in range(T):
        for s_idx in range(N_STATES):
            style_formula(ws, 5, fc + t_idx * 2 + s_idx, formulas_fwd[t_idx][s_idx])

    # Row 6: values
    for t_idx in range(T):
        for s_idx in range(N_STATES):
            write_val(ws, 6, fc + t_idx * 2 + s_idx, alpha[t_idx, s_idx])

    # P(O|λ)
    r_po = 8
    style_subheader(ws, r_po, fc, "P(O|λ) = Σ α_T(i)")
    ws.merge_cells(start_row=r_po, start_column=fc, end_row=r_po, end_column=fc + 2)
    write_val(ws, r_po, fc + 3, p_obs)

    # ---- Section C: Backward (cols 16-21) ----
    bc = 16
    style_header(ws, 1, bc, "Backward Algorithm (β)", span=6)
    # Steps go t=3,2,1
    step_labels_bw = ["Step 1: Init (t=3)", "Step 2: Recursion (t=2)", "Step 3: Recursion (t=1)"]
    t_order_bw = [2, 1, 0]
    for idx in range(3):
        style_header(ws, 2, bc + idx * 2, step_labels_bw[idx], span=2, fill=LIGHT_BLUE, font=BOLD)
    for idx, t_idx in enumerate(t_order_bw):
        obs_sym = OBS_LABELS[t_idx]
        style_subheader(ws, 3, bc + idx * 2, f"t={t_idx+1}")
        ws.merge_cells(start_row=3, start_column=bc + idx * 2,
                       end_row=3, end_column=bc + idx * 2 + 1)
    for idx, t_idx in enumerate(t_order_bw):
        for s_idx, s in enumerate(["R", "S"]):
            style_subheader(ws, 4, bc + idx * 2 + s_idx, f"β{t_idx+1}({s})")

    formulas_bw = [
        ["1 (base case)", "1 (base case)"],
        ["Σ_j a_Rj·B_j(o3)·β3(j)", "Σ_j a_Sj·B_j(o3)·β3(j)"],
        ["Σ_j a_Rj·B_j(o2)·β2(j)", "Σ_j a_Sj·B_j(o2)·β2(j)"],
    ]
    for idx in range(3):
        for s_idx in range(N_STATES):
            style_formula(ws, 5, bc + idx * 2 + s_idx, formulas_bw[idx][s_idx])

    for idx, t_idx in enumerate(t_order_bw):
        for s_idx in range(N_STATES):
            write_val(ws, 6, bc + idx * 2 + s_idx, beta[t_idx, s_idx])

    # Consistency check
    r_cc = 8
    style_subheader(ws, r_cc, bc, "Consistency: Σ α_t·β_t (each t)")
    ws.merge_cells(start_row=r_cc, start_column=bc, end_row=r_cc, end_column=bc + 3)
    for t_idx in range(T):
        val = (alpha[t_idx] * beta[t_idx]).sum()
        write_val(ws, r_cc, bc + 4 + t_idx, val)

    # ---- Section D: Posterior γ (cols 23-28) ----
    gc = 23
    style_header(ws, 1, gc, "Posterior Probabilities γ", span=6)
    style_subheader(ws, 2, gc, "γ_t(i) = α_t(i)·β_t(i) / P(O|λ)")
    ws.merge_cells(start_row=2, start_column=gc, end_row=2, end_column=gc + 5)

    for t_idx in range(T):
        style_subheader(ws, 3, gc + t_idx * 2, f"t={t_idx+1}")
        ws.merge_cells(start_row=3, start_column=gc + t_idx * 2,
                       end_row=3, end_column=gc + t_idx * 2 + 1)
    for t_idx in range(T):
        for s_idx, s in enumerate(["R", "S"]):
            style_subheader(ws, 4, gc + t_idx * 2 + s_idx, f"γ{t_idx+1}({s})")

    for t_idx in range(T):
        for s_idx in range(N_STATES):
            write_val(ws, 5, gc + t_idx * 2 + s_idx, gamma[t_idx, s_idx])

    # Most likely state
    r_ml = 7
    style_subheader(ws, r_ml, gc, "Most Likely State")
    ws.merge_cells(start_row=r_ml, start_column=gc, end_row=r_ml, end_column=gc + 1)
    for t_idx in range(T):
        best = STATES[int(np.argmax(gamma[t_idx]))]
        write_text(ws, r_ml, gc + 2 + t_idx, best, bold=True)

    auto_width(ws, min_width=12, max_width=30)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Sheet 3: Baum-Welch
# ---------------------------------------------------------------------------

def build_baum_welch_sheet(wb):
    ws = wb.create_sheet("Baum-Welch")

    # Run Baum-Welch iterations collecting results
    pi_cur = PI.copy()
    a_cur = A.copy()
    b_cur = B.copy()
    obs = OBS_SEQ

    iterations = []
    for it in range(BW_ITERATIONS):
        pi_new, a_new, b_new, alpha, beta, gamma, xi, ll = baum_welch_step(
            pi_cur, a_cur, b_cur, obs
        )
        iterations.append({
            "iter": it,
            "pi_in": pi_cur.copy(), "a_in": a_cur.copy(), "b_in": b_cur.copy(),
            "alpha": alpha.copy(), "beta": beta.copy(),
            "gamma": gamma.copy(), "xi": xi.copy(),
            "pi_out": pi_new.copy(), "a_out": a_new.copy(), "b_out": b_new.copy(),
            "ll": ll,
        })
        pi_cur, a_cur, b_cur = pi_new, a_new, b_new

    # --- Build the sheet ---
    # Column layout:
    # 1: Iteration
    # 2: Log-likelihood
    # 3-4: π(R), π(S)
    # 5-8: A: a_RR, a_RS, a_SR, a_SS
    # 9-12: B: b_R(U), b_R(N), b_S(U), b_S(N)
    # 13-16: γ1(R), γ1(S), γ2(R), γ2(S)  (γ3 omitted for brevity)
    # 17-18: γ3(R), γ3(S)
    # 19-22: ξ1(RR), ξ1(RS), ξ1(SR), ξ1(SS)
    # 23-26: ξ2(RR), ξ2(RS), ξ2(SR), ξ2(SS)

    # Row 1: major headers
    style_header(ws, 1, 1, "Baum-Welch EM Iterations", span=26)

    # Row 2: section headers
    style_header(ws, 2, 1, "Iter", fill=DARK_BLUE)
    style_header(ws, 2, 2, "Log P(O|λ)", fill=DARK_BLUE)
    style_header(ws, 2, 3, "Re-estimated π", span=2, fill=DARK_BLUE)
    style_header(ws, 2, 5, "Re-estimated A", span=4, fill=DARK_BLUE)
    style_header(ws, 2, 9, "Re-estimated B", span=4, fill=DARK_BLUE)
    style_header(ws, 2, 13, "γ (t=1,2,3)", span=6, fill=DARK_BLUE)
    style_header(ws, 2, 19, "ξ (t=1)", span=4, fill=DARK_BLUE)
    style_header(ws, 2, 23, "ξ (t=2)", span=4, fill=DARK_BLUE)

    # Row 3: sub-headers
    sub = [
        (1, "Iter"), (2, "ln P(O|λ)"),
        (3, "π(R)"), (4, "π(S)"),
        (5, "a(R,R)"), (6, "a(R,S)"), (7, "a(S,R)"), (8, "a(S,S)"),
        (9, "b_R(U)"), (10, "b_R(N)"), (11, "b_S(U)"), (12, "b_S(N)"),
        (13, "γ1(R)"), (14, "γ1(S)"), (15, "γ2(R)"), (16, "γ2(S)"),
        (17, "γ3(R)"), (18, "γ3(S)"),
        (19, "ξ1(R,R)"), (20, "ξ1(R,S)"), (21, "ξ1(S,R)"), (22, "ξ1(S,S)"),
        (23, "ξ2(R,R)"), (24, "ξ2(R,S)"), (25, "ξ2(S,R)"), (26, "ξ2(S,S)"),
    ]
    for col, label in sub:
        style_subheader(ws, 3, col, label)

    # Row 4: formula descriptions
    formula_descs = {
        1: "—",
        2: "ln Σ α_T(i)",
        3: "γ1(R)", 4: "γ1(S)",
        5: "Σ_t ξ_t(R,R)/Σ_t γ_t(R)", 6: "Σ_t ξ_t(R,S)/Σ_t γ_t(R)",
        7: "Σ_t ξ_t(S,R)/Σ_t γ_t(S)", 8: "Σ_t ξ_t(S,S)/Σ_t γ_t(S)",
        9: "Σ_{t:o=U} γ_t(R)/Σ_t γ_t(R)", 10: "Σ_{t:o=N} γ_t(R)/Σ_t γ_t(R)",
        11: "Σ_{t:o=U} γ_t(S)/Σ_t γ_t(S)", 12: "Σ_{t:o=N} γ_t(S)/Σ_t γ_t(S)",
    }
    for col, desc in formula_descs.items():
        style_formula(ws, 4, col, desc)

    # Rows 5+: data
    # Row 5 = initial params (iteration 0 input)
    data_start = 5
    for idx, it_data in enumerate(iterations):
        r = data_start + idx
        write_val(ws, r, 1, idx, fmt="0")
        write_val(ws, r, 2, it_data["ll"])
        # re-estimated pi
        write_val(ws, r, 3, it_data["pi_out"][0])
        write_val(ws, r, 4, it_data["pi_out"][1])
        # re-estimated A
        write_val(ws, r, 5, it_data["a_out"][0, 0])
        write_val(ws, r, 6, it_data["a_out"][0, 1])
        write_val(ws, r, 7, it_data["a_out"][1, 0])
        write_val(ws, r, 8, it_data["a_out"][1, 1])
        # re-estimated B
        write_val(ws, r, 9, it_data["b_out"][0, 0])
        write_val(ws, r, 10, it_data["b_out"][0, 1])
        write_val(ws, r, 11, it_data["b_out"][1, 0])
        write_val(ws, r, 12, it_data["b_out"][1, 1])
        # gamma
        write_val(ws, r, 13, it_data["gamma"][0, 0])
        write_val(ws, r, 14, it_data["gamma"][0, 1])
        write_val(ws, r, 15, it_data["gamma"][1, 0])
        write_val(ws, r, 16, it_data["gamma"][1, 1])
        write_val(ws, r, 17, it_data["gamma"][2, 0])
        write_val(ws, r, 18, it_data["gamma"][2, 1])
        # xi t=1
        write_val(ws, r, 19, it_data["xi"][0, 0, 0])
        write_val(ws, r, 20, it_data["xi"][0, 0, 1])
        write_val(ws, r, 21, it_data["xi"][0, 1, 0])
        write_val(ws, r, 22, it_data["xi"][0, 1, 1])
        # xi t=2
        write_val(ws, r, 23, it_data["xi"][1, 0, 0])
        write_val(ws, r, 24, it_data["xi"][1, 0, 1])
        write_val(ws, r, 25, it_data["xi"][1, 1, 0])
        write_val(ws, r, 26, it_data["xi"][1, 1, 1])

    auto_width(ws, min_width=12, max_width=24)
    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    build_parameters_sheet(wb)
    build_forward_backward_sheet(wb)
    build_baum_welch_sheet(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "docs")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "HMM_Forward_Backward_Computation.xlsx")
    wb.save(out_path)
    print(f"Saved: {os.path.abspath(out_path)}")


if __name__ == "__main__":
    main()
